import sys
import json
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import os

def generate_report(data_file, output_path):
    with open(data_file, 'r') as f:
        data = json.load(f)
    
    wb = Workbook()
    ws1 = wb.active
    report_type = data.get('report_type', 'Inventory Report')
    ws1.title = report_type[:30]

    # --- STYLING ---
    green_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, bold=True, color="059669")
    header_font = Font(size=14, bold=True, color="4b5563")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- HEADER & LOGO ---
    # Add Logo if exists
    if os.path.exists('library_logo.png'):
        img = XLImage('library_logo.png')
        img.width = 60
        img.height = 60
        ws1.add_image(img, 'B2')

    ws1["C2"] = "SUMMIT KNOWLEDGE LIBRARY"
    ws1["C2"].font = title_font
    ws1["C3"] = f"System Generated {report_type}"
    ws1["C3"].font = header_font

    # --- TABLE HEADERS ---
    headers = data['headers']
    start_row = 7
    for col_num, header in enumerate(headers, 2):
        cell = ws1.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border

    # --- TABLE DATA ---
    row_num = start_row + 1
    for row_data in data['rows']:
        for col_idx, value in enumerate(row_data, 2):
            cell = ws1.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            # Status coloring
            if headers[col_idx-2] == "Status":
                if value == "Available" or value == "Returned":
                    cell.font = Font(color="059669", bold=True)
                elif value == "Borrowed" or value == "Overdue":
                    cell.font = Font(color="dc2626", bold=True)
        row_num += 1

    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws1.column_dimensions[column].width = max_length + 5

    # --- SIGNATURE PLACEHOLDER ---
    sig_row = row_num + 3
    ws1.cell(row=sig_row, column=2, value="Prepared By:")
    ws1.cell(row=sig_row+1, column=2, value="____________________")
    ws1.cell(row=sig_row+2, column=2, value="System Administrator")

    # --- SHEET 2: CHART ---
    ws2 = wb.create_sheet(title="Chart Report")
    
    # Chart Data Table
    ws2["A1"] = "DATA ANALYSIS SUMMARY"
    ws2["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws2["A1"].fill = green_fill
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = center_align

    ws2.cell(row=3, column=1, value="Category").fill = green_fill
    ws2.cell(row=3, column=1).font = white_font
    ws2.cell(row=3, column=2, value="Count").fill = green_fill
    ws2.cell(row=3, column=2).font = white_font

    chart_data = data['chart_data']
    for idx, (label, val) in enumerate(chart_data.items(), 4):
        ws2.cell(row=idx, column=1, value=label).border = thin_border
        ws2.cell(row=idx, column=2, value=val).border = thin_border

    # PIE CHART
    pie = PieChart()
    labels = Reference(ws2, min_col=1, min_row=4, max_row=4 + len(chart_data) - 1)
    data_ref = Reference(ws2, min_col=2, min_row=3, max_row=4 + len(chart_data) - 1)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = f"{report_type} Distribution"
    pie.style = 10
    
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    
    ws2.add_chart(pie, "D2")

    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) > 2:
        generate_report(sys.argv[1], sys.argv[2])
